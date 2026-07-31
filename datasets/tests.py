import json
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .equivalences import (
    analyze_numeric_columns,
    transform_records,
    validate_equivalence_payload,
)
from .models import (
    Dataset,
    DatasetEquivalenceApplication,
    EquivalenceConfiguration,
)
from .services import filter_dataset_by_category


class DatasetViewTests(TestCase):
    def test_user_can_select_any_low_cardinality_filter_column(self):
        dataset = Dataset.objects.create(
            pk=1, source_name='generico.csv',
            columns=['Grupo', 'valor'],
            records=[
                {'Grupo': 'A' if index < 3 else 'B', 'valor': str(index)}
                for index in range(6)
            ],
        )

        filtered = filter_dataset_by_category(dataset, 'a', 'Grupo')

        self.assertEqual(filtered['category_column'], 'Grupo')
        self.assertEqual(len(filtered['filtered_records']), 3)
        self.assertEqual(
            [item['name'] for item in filtered['category_columns']],
            ['Grupo'],
        )
        response = self.client.get(
            reverse('dashboard:index'),
            {'category_column': 'Grupo', 'category': 'a'},
        )
        self.assertContains(response, 'Columna: Grupo')
        self.assertContains(response, 'category_column=Grupo')

    def test_csv_upload_persists_columns_and_records(self):
        csv_file = SimpleUploadedFile(
            'personas.csv',
            b'nombre,edad\nAna,28\nLuis,34\n',
            content_type='text/csv',
        )

        response = self.client.post(reverse('datasets:upload'), {'file': csv_file})

        self.assertRedirects(response, reverse('dashboard:index'))
        dataset = Dataset.objects.get(pk=1)
        self.assertEqual(dataset.columns, ['nombre', 'edad'])
        self.assertEqual(dataset.records[0], {'nombre': 'Ana', 'edad': '28'})

    def test_excel_upload_persists_columns_and_records(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['nombre', 'edad', 'categoria'])
        worksheet.append(['Ana', 28, 'Agua'])
        worksheet.append(['Luis', 34, 'Tierra'])
        content = BytesIO()
        workbook.save(content)
        excel_file = SimpleUploadedFile(
            'personas.xlsx',
            content.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )

        response = self.client.post(
            reverse('datasets:upload'), {'file': excel_file}
        )

        self.assertRedirects(response, reverse('dashboard:index'))
        dataset = Dataset.objects.get(pk=1)
        self.assertEqual(dataset.source_name, 'personas.xlsx')
        self.assertEqual(dataset.columns, ['nombre', 'edad', 'categoria'])
        self.assertEqual(
            dataset.records[0],
            {'nombre': 'Ana', 'edad': '28', 'categoria': 'Agua'},
        )

    def test_invalid_excel_does_not_replace_current_dataset(self):
        original = Dataset.objects.create(
            pk=1,
            source_name='actual.csv',
            columns=['valor'],
            records=[{'valor': '1'}],
        )
        invalid_excel = SimpleUploadedFile(
            'datos.xlsx',
            b'este contenido no es un libro de Excel',
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )

        response = self.client.post(
            reverse('datasets:upload'), {'file': invalid_excel}
        )

        self.assertRedirects(
            response, f"{reverse('dashboard:index')}?upload=invalid"
        )
        original.refresh_from_db()
        self.assertEqual(original.source_name, 'actual.csv')

    def test_new_upload_replaces_previous_dataset(self):
        Dataset.objects.create(
            pk=1,
            source_name='anterior.csv',
            columns=['anterior'],
            records=[{'anterior': 'dato'}],
        )
        new_file = SimpleUploadedFile(
            'nuevo.csv', b'categoria,valor\nA,10\n', content_type='text/csv'
        )

        self.client.post(reverse('datasets:upload'), {'file': new_file})

        self.assertEqual(Dataset.objects.count(), 1)
        dataset = Dataset.objects.get(pk=1)
        self.assertEqual(dataset.source_name, 'nuevo.csv')
        self.assertEqual(dataset.columns, ['categoria', 'valor'])

    def test_invalid_file_does_not_replace_current_dataset(self):
        original = Dataset.objects.create(
            pk=1,
            source_name='actual.csv',
            columns=['valor'],
            records=[{'valor': '1'}],
        )
        invalid_file = SimpleUploadedFile(
            'datos.txt', b'valor\n2\n', content_type='text/plain'
        )

        response = self.client.post(
            reverse('datasets:upload'), {'file': invalid_file}
        )

        self.assertRedirects(
            response, f"{reverse('dashboard:index')}?upload=invalid"
        )
        original.refresh_from_db()
        self.assertEqual(original.source_name, 'actual.csv')

    def test_single_column_csv_is_accepted(self):
        csv_file = SimpleUploadedFile(
            'valores.csv', b'valor\nuno\ndos\n', content_type='text/csv'
        )

        response = self.client.post(reverse('datasets:upload'), {'file': csv_file})

        self.assertRedirects(response, reverse('dashboard:index'))
        self.assertEqual(Dataset.objects.get().row_count, 2)

    def test_dataset_table_paginates_twenty_five_rows(self):
        self.create_numbered_dataset(60)

        response = self.client.get(reverse('dashboard:index'), {'page': 2})

        self.assertEqual(response.context['page_obj'].number, 2)
        self.assertEqual(len(response.context['table_rows']), 25)
        self.assertEqual(response.context['table_rows'][0]['number'], 26)
        self.assertEqual(response.context['table_rows'][-1]['number'], 50)

    def test_out_of_range_page_returns_last_page(self):
        self.create_numbered_dataset(30)

        response = self.client.get(reverse('dashboard:index'), {'page': 99})

        self.assertEqual(response.context['page_obj'].number, 2)
        self.assertEqual(len(response.context['table_rows']), 5)

    def test_dataset_can_be_removed_manually(self):
        self.create_numbered_dataset(1)

        response = self.client.post(reverse('datasets:delete'))

        self.assertRedirects(response, reverse('dashboard:index'))
        self.assertFalse(Dataset.objects.exists())

    def test_pagination_links_return_to_dataset_table(self):
        self.create_numbered_dataset(30)

        response = self.client.get(reverse('dashboard:index'))

        self.assertContains(response, '?representation=original&amp;page=2#dataset-table')

    def test_categories_are_grouped_without_case_sensitivity(self):
        Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['Categoria', 'nombre'],
            records=[
                {'Categoria': 'Agua', 'nombre': 'Uno'},
                {'Categoria': 'AGUA', 'nombre': 'Dos'},
                {'Categoria': ' agua ', 'nombre': 'Tres'},
                {'Categoria': 'TIERRA', 'nombre': 'Cuatro'},
            ],
        )

        response = self.client.get(reverse('dashboard:index'))

        self.assertEqual(
            response.context['category_options'],
            [
                {'value': 'agua', 'label': 'Agua'},
                {'value': 'tierra', 'label': 'Tierra'},
            ],
        )

    def test_category_filter_returns_all_matching_case_variants(self):
        Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria', 'nombre'],
            records=[
                {'categoria': 'Agua', 'nombre': 'Uno'},
                {'categoria': 'AGUA', 'nombre': 'Dos'},
                {'categoria': 'agua', 'nombre': 'Tres'},
                {'categoria': 'Tierra', 'nombre': 'Cuatro'},
            ],
        )

        response = self.client.get(
            reverse('dashboard:index'), {'category': 'AGUA'}
        )

        self.assertEqual(response.context['selected_category'], 'agua')
        self.assertEqual(response.context['page_obj'].paginator.count, 3)
        self.assertEqual(
            [row['values'][1] for row in response.context['table_rows']],
            ['Uno', 'Dos', 'Tres'],
        )

    def test_unknown_category_falls_back_to_all_records(self):
        Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria'],
            records=[{'categoria': 'Agua'}, {'categoria': 'Tierra'}],
        )

        response = self.client.get(
            reverse('dashboard:index'), {'category': 'inexistente'}
        )

        self.assertEqual(response.context['selected_category'], '')
        self.assertEqual(response.context['page_obj'].paginator.count, 2)

    def test_pagination_keeps_selected_category(self):
        Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria', 'valor'],
            records=[
                {'categoria': 'Agua', 'valor': str(number)}
                for number in range(30)
            ],
        )

        response = self.client.get(
            reverse('dashboard:index'), {'category': 'agua'}
        )

        self.assertContains(
            response,
            '?category=agua&amp;representation=original&amp;page=2#dataset-table',
        )

    def test_download_button_only_appears_with_an_active_filter(self):
        Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria'],
            records=[{'categoria': 'Agua'}, {'categoria': 'Tierra'}],
        )

        unfiltered_response = self.client.get(reverse('dashboard:index'))
        filtered_response = self.client.get(
            reverse('dashboard:index'), {'category': 'agua'}
        )

        self.assertNotContains(unfiltered_response, 'Descargar Excel')
        self.assertContains(filtered_response, 'Descargar Excel')
        self.assertContains(
            filtered_response,
            '/datasets/descargar/?category=agua&amp;representation=original',
        )

    def test_download_button_uses_the_representation_visible_on_screen(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria', 'Escala'],
            records=[{'categoria': 'Agua', 'Escala': '1'}],
        )
        configuration = EquivalenceConfiguration.objects.create(
            name='Escala',
            mapping={'1': 'Bajo'},
            possible_values=['1'],
        )
        DatasetEquivalenceApplication.objects.create(
            dataset=dataset,
            configuration=configuration,
            columns=['Escala'],
        )

        response = self.client.get(
            reverse('dashboard:index'),
            {'category': 'agua', 'representation': 'qualitative'},
        )

        self.assertContains(
            response,
            '/datasets/descargar/?category=agua&amp;representation=qualitative',
        )

    def test_excel_download_contains_every_filtered_case_variant(self):
        Dataset.objects.create(
            pk=1,
            source_name='Datos Generales.csv',
            columns=['nombre', 'categoria'],
            records=[
                {'nombre': 'Uno', 'categoria': 'Agua'},
                {'nombre': 'Dos', 'categoria': 'AGUA'},
                {'nombre': 'Tres', 'categoria': 'Tierra'},
            ],
        )

        response = self.client.get(
            reverse('datasets:download'), {'category': 'agua'}
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Datos filtrados']
        rows = [
            list(row)
            for row in worksheet.iter_rows(values_only=True)
        ]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            'datos-generales-agua-original.xlsx',
            response['Content-Disposition'],
        )
        self.assertEqual(worksheet.freeze_panes, 'A2')
        self.assertIn('DatasetFiltrado', worksheet.tables)
        self.assertEqual(
            rows,
            [
                ['nombre', 'categoria'],
                ['Uno', 'Agua'],
                ['Dos', 'AGUA'],
            ],
        )

    def test_download_without_a_valid_filter_is_not_available(self):
        Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria'],
            records=[{'categoria': 'Agua'}],
        )

        response = self.client.get(reverse('datasets:download'))

        self.assertEqual(response.status_code, 404)

    def test_scale_columns_with_partial_observed_values_are_compatible(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='escalas.csv',
            columns=['Columna A', 'Columna B'],
            records=[
                {'Columna A': str(value), 'Columna B': str(max(value, 2))}
                for value in range(1, 6)
            ],
        )
        payload = {
            'name': 'Escala del uno al cinco',
            'equivalences': [
                {'value': str(value), 'label': f'Nivel {value}'}
                for value in range(1, 6)
            ],
            'columns': ['Columna A', 'Columna B'],
        }

        cleaned, errors = validate_equivalence_payload(dataset, payload)

        self.assertEqual(errors, {})
        self.assertEqual(cleaned['columns'], ['Columna A', 'Columna B'])

    def test_unconfigured_observed_value_marks_column_incompatible(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='escala.csv',
            columns=['Escala'],
            records=[{'Escala': str(value)} for value in range(1, 7)],
        )
        payload = {
            'name': 'Escala incompleta',
            'equivalences': [
                {'value': str(value), 'label': f'Nivel {value}'}
                for value in range(1, 6)
            ],
            'columns': ['Escala'],
        }

        _, errors = validate_equivalence_payload(dataset, payload)

        self.assertEqual(errors['incompatible_columns'], {'Escala': ['6']})

    def test_binary_configuration_converts_zero_and_one(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='binario.csv',
            columns=['Participa'],
            records=[{'Participa': '0'}, {'Participa': '1'}],
        )
        configuration = EquivalenceConfiguration.objects.create(
            name='Sí o no',
            mapping={'0': 'No', '1': 'Sí'},
            possible_values=['0', '1'],
        )
        DatasetEquivalenceApplication.objects.create(
            dataset=dataset,
            configuration=configuration,
            columns=['Participa'],
        )

        columns, records, representation = transform_records(
            dataset, dataset.records, 'qualitative'
        )

        self.assertEqual(columns, ['Participa'])
        self.assertEqual(records, [{'Participa': 'No'}, {'Participa': 'Sí'}])
        self.assertEqual(representation, 'qualitative')

    def test_continuous_age_is_numeric_but_not_recommended(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='edades.csv',
            columns=['Edad'],
            records=[{'Edad': str(age)} for age in range(18, 61)],
        )

        analysis = analyze_numeric_columns(dataset)[0]

        self.assertEqual(analysis['name'], 'Edad')
        self.assertFalse(analysis['recommended'])

    def test_numeric_identifier_is_not_recommended(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='alumnos.csv',
            columns=['Matrícula'],
            records=[{'Matrícula': str(value)} for value in range(1001, 1007)],
        )

        analysis = analyze_numeric_columns(dataset)[0]

        self.assertTrue(analysis['possible_identifier'])
        self.assertFalse(analysis['recommended'])

    def test_null_values_are_ignored_and_preserved_during_conversion(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='nulos.csv',
            columns=['Escala'],
            records=[
                {'Escala': '1'},
                {'Escala': '2'},
                {'Escala': ''},
                {'Escala': '4'},
                {'Escala': 'NaN'},
            ],
        )
        analysis = analyze_numeric_columns(dataset)[0]
        configuration = EquivalenceConfiguration.objects.create(
            name='Escala con nulos',
            mapping={'1': 'Bajo', '2': 'Medio', '4': 'Alto'},
            possible_values=['1', '2', '4'],
        )
        DatasetEquivalenceApplication.objects.create(
            dataset=dataset,
            configuration=configuration,
            columns=['Escala'],
        )

        _, records, _ = transform_records(dataset, dataset.records, 'qualitative')

        self.assertEqual(analysis['unique_values'], ['1', '2', '4'])
        self.assertEqual(analysis['null_count'], 2)
        self.assertEqual(records[2]['Escala'], '')
        self.assertEqual(records[4]['Escala'], 'NaN')

    def test_more_than_fifty_numeric_columns_are_available_to_modal(self):
        columns = [f'Escala {index}' for index in range(55)]
        dataset = Dataset.objects.create(
            pk=1,
            source_name='amplio.csv',
            columns=columns,
            records=[{column: '1' for column in columns}],
        )

        response = self.client.get(reverse('dashboard:index'))

        self.assertEqual(len(response.context['equivalence_data']['numeric_columns']), 55)
        self.assertContains(response, 'style="max-height: 16rem;"')

    def test_save_endpoint_persists_configuration_and_application(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='encuesta.csv',
            columns=['Pregunta 1', 'Pregunta 2'],
            records=[
                {'Pregunta 1': '1', 'Pregunta 2': '2'},
                {'Pregunta 1': '2', 'Pregunta 2': '1'},
            ],
        )
        payload = {
            'name': 'Escala binaria',
            'equivalences': [
                {'value': '1', 'label': 'No'},
                {'value': '2', 'label': 'Sí'},
            ],
            'columns': dataset.columns,
        }

        response = self.client.post(
            reverse('datasets:equivalence_save'),
            data=json.dumps(payload),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        configuration = EquivalenceConfiguration.objects.get()
        application = DatasetEquivalenceApplication.objects.get()
        self.assertEqual(configuration.mapping, {'1': 'No', '2': 'Sí'})
        self.assertEqual(application.columns, dataset.columns)

    def test_save_endpoint_rejects_blank_labels_duplicates_and_no_columns(self):
        Dataset.objects.create(
            pk=1,
            source_name='encuesta.csv',
            columns=['Escala'],
            records=[{'Escala': '1'}],
        )
        payload = {
            'name': '',
            'equivalences': [
                {'value': '1', 'label': ''},
                {'value': '1.0', 'label': 'Repetido'},
            ],
            'columns': [],
        }

        response = self.client.post(
            reverse('datasets:equivalence_save'),
            data=json.dumps(payload),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        errors = response.json()['errors']
        self.assertIn('name', errors)
        self.assertIn('equivalence_rows', errors)
        self.assertIn('columns', errors)

    def test_replacing_dataset_keeps_template_but_removes_application(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='anterior.csv',
            columns=['Escala'],
            records=[{'Escala': '1'}],
        )
        configuration = EquivalenceConfiguration.objects.create(
            name='Plantilla reutilizable',
            mapping={'1': 'Sí'},
            possible_values=['1'],
        )
        DatasetEquivalenceApplication.objects.create(
            dataset=dataset,
            configuration=configuration,
            columns=['Escala'],
        )
        new_file = SimpleUploadedFile(
            'nuevo.csv', b'Escala\n1\n', content_type='text/csv'
        )

        self.client.post(reverse('datasets:upload'), {'file': new_file})

        self.assertTrue(EquivalenceConfiguration.objects.filter(pk=configuration.pk).exists())
        self.assertFalse(DatasetEquivalenceApplication.objects.exists())

    def test_filter_is_applied_before_qualitative_download(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='encuesta.csv',
            columns=['categoria', 'Escala'],
            records=[
                {'categoria': 'Agua', 'Escala': '1'},
                {'categoria': 'Agua', 'Escala': '2'},
                {'categoria': 'Tierra', 'Escala': '1'},
            ],
        )
        configuration = EquivalenceConfiguration.objects.create(
            name='Escala',
            mapping={'1': 'Bajo', '2': 'Alto'},
            possible_values=['1', '2'],
        )
        DatasetEquivalenceApplication.objects.create(
            dataset=dataset,
            configuration=configuration,
            columns=['Escala'],
        )

        response = self.client.get(
            reverse('datasets:download'),
            {'category': 'agua', 'representation': 'qualitative'},
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = [
            list(row)
            for row in workbook['Datos filtrados'].iter_rows(values_only=True)
        ]

        self.assertEqual(
            rows,
            [
                ['categoria', 'Escala'],
                ['Agua', 'Bajo'],
                ['Agua', 'Alto'],
            ],
        )

    def test_configuration_can_be_removed_from_dataset_and_reused(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='encuesta.csv',
            columns=['Escala'],
            records=[{'Escala': '1'}],
        )
        configuration = EquivalenceConfiguration.objects.create(
            name='Plantilla',
            mapping={'1': 'Sí'},
            possible_values=['1'],
        )
        DatasetEquivalenceApplication.objects.create(
            dataset=dataset,
            configuration=configuration,
            columns=['Escala'],
        )

        response = self.client.post(
            reverse(
                'datasets:equivalence_remove_application',
                args=[configuration.pk],
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            EquivalenceConfiguration.objects.filter(pk=configuration.pk).exists()
        )
        self.assertFalse(DatasetEquivalenceApplication.objects.exists())

    def test_deleting_configuration_also_deletes_its_application(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='encuesta.csv',
            columns=['Escala'],
            records=[{'Escala': '1'}],
        )
        configuration = EquivalenceConfiguration.objects.create(
            name='Temporal',
            mapping={'1': 'Sí'},
            possible_values=['1'],
        )
        DatasetEquivalenceApplication.objects.create(
            dataset=dataset,
            configuration=configuration,
            columns=['Escala'],
        )

        response = self.client.post(
            reverse('datasets:equivalence_delete', args=[configuration.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(EquivalenceConfiguration.objects.exists())
        self.assertFalse(DatasetEquivalenceApplication.objects.exists())

    @staticmethod
    def create_numbered_dataset(row_count):
        return Dataset.objects.create(
            pk=1,
            source_name='datos.csv',
            columns=['valor'],
            records=[{'valor': str(number)} for number in range(row_count)],
        )
