from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def build_excel_file_from_rows(
    columns,
    rows,
    *,
    worksheet_title='Datos filtrados',
    table_name='DatasetFiltrado',
):
    """Create a formatted Excel workbook from ordered headers and row values."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title
    worksheet.freeze_panes = 'A2'

    worksheet.append(columns)
    for row in rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(
                ('=', '+', '-', '@')
            ):
                cell.data_type = 's'

    if worksheet.max_row > 1 and worksheet.max_column:
        table_reference = (
            f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'
        )
        table = Table(displayName=table_name, ref=table_reference)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    for index, column in enumerate(columns, start=1):
        values = [column] + [
            str(row[index - 1] if index <= len(row) else '')
            for row in rows[:200]
        ]
        width = min(max(len(value) for value in values) + 2, 40)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_excel_file(columns, records):
    """Create a formatted Excel workbook without modifying source records."""
    rows = [
        [record.get(column, '') for column in columns]
        for record in records
    ]
    return build_excel_file_from_rows(columns, rows)
