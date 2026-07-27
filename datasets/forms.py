import csv
import io
from datetime import date, datetime, time
from pathlib import Path
from zipfile import BadZipFile

from django import forms
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


MAX_FILE_SIZE = 10 * 1024 * 1024
SUPPORTED_EXTENSIONS = {'.csv', '.xlsx'}


def _normalize_cell(value):
    if value is None:
        return ''
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value).strip()


def _normalize_dataset(raw_columns, raw_rows):
    columns = [_normalize_cell(column) for column in raw_columns]
    while columns and not columns[-1]:
        columns.pop()

    if not columns or any(not column for column in columns):
        raise forms.ValidationError('El archivo debe incluir encabezados válidos.')
    if len(columns) != len(set(columns)):
        raise forms.ValidationError('Los nombres de las columnas no pueden repetirse.')

    records = []
    for raw_row in raw_rows:
        row = list(raw_row)
        if len(row) > len(columns) and any(
            _normalize_cell(value) for value in row[len(columns):]
        ):
            raise forms.ValidationError(
                'Hay registros con más valores que columnas.'
            )
        values = [
            _normalize_cell(row[index]) if index < len(row) else ''
            for index in range(len(columns))
        ]
        record = dict(zip(columns, values))
        if any(record.values()):
            records.append(record)

    if not records:
        raise forms.ValidationError('El archivo no contiene registros de datos.')
    return columns, records


def _read_csv(uploaded_file):
    try:
        content = uploaded_file.read().decode('utf-8-sig')
    except UnicodeDecodeError as error:
        raise forms.ValidationError(
            'El archivo CSV debe utilizar codificación UTF-8.'
        ) from error

    try:
        dialect = csv.Sniffer().sniff(content[:4096], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel

    try:
        reader = csv.reader(io.StringIO(content), dialect=dialect)
        columns = next(reader, [])
        return _normalize_dataset(columns, reader)
    except csv.Error as error:
        raise forms.ValidationError(
            'No fue posible interpretar el archivo CSV.'
        ) from error


def _read_excel(uploaded_file):
    uploaded_file.seek(0)
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        columns = next(rows, [])
        return _normalize_dataset(columns, rows)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise forms.ValidationError(
            'No fue posible interpretar el archivo Excel.'
        ) from error
    finally:
        workbook_to_close = locals().get('workbook')
        if workbook_to_close:
            workbook_to_close.close()


class DatasetUploadForm(forms.Form):
    file = forms.FileField(
        label='Archivo de datos',
        widget=forms.ClearableFileInput(
            attrs={'class': 'form-control', 'accept': '.csv,.xlsx'}
        ),
    )

    def clean_file(self):
        uploaded_file = self.cleaned_data['file']
        extension = Path(uploaded_file.name).suffix.lower()

        if extension not in SUPPORTED_EXTENSIONS:
            raise forms.ValidationError(
                'Selecciona un archivo con formato CSV o XLSX.'
            )
        if len(uploaded_file.name) > 255:
            raise forms.ValidationError('El nombre del archivo es demasiado largo.')
        if uploaded_file.size > MAX_FILE_SIZE:
            raise forms.ValidationError('El archivo no debe superar los 10 MB.')
        return uploaded_file

    def clean(self):
        cleaned_data = super().clean()
        uploaded_file = cleaned_data.get('file')
        if not uploaded_file:
            return cleaned_data

        extension = Path(uploaded_file.name).suffix.lower()
        if extension == '.xlsx':
            columns, records = _read_excel(uploaded_file)
        else:
            columns, records = _read_csv(uploaded_file)

        cleaned_data['columns'] = columns
        cleaned_data['records'] = records
        return cleaned_data
