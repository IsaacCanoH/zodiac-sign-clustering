from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from datasets.models import Dataset
from dbscan.services import train_dbscan
from kmeans.services import train_kmeans

from .services import PAGE_SIZE, build_classified_context


class ClassifiedDataTests(TestCase):
    def test_empty_state_explains_that_training_is_required(self):
        Dataset.objects.create(
            pk=1,
            source_name='datos.csv',
            columns=['x'],
            records=[{'x': '1'}, {'x': '2'}],
        )

        response = self.client.get(reverse('dashboard:index'))

        self.assertContains(response, 'Datos clasificados')
        self.assertContains(response, 'Todavía no existen datos clasificados')
        self.assertContains(response, 'Ejecuta un entrenamiento de K-Means o DBSCAN')

    def test_kmeans_shows_original_columns_row_and_cluster(self):
        dataset = self.create_dataset(30)
        run = train_kmeans(dataset, ['x', 'y'], 2)

        response = self.client.get(
            reverse('dashboard:index'),
            {'classified_algorithm': 'kmeans'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Fila original')
        self.assertContains(response, 'nombre')
        self.assertContains(response, 'x')
        self.assertContains(response, 'y')
        self.assertContains(response, 'Cluster K-Means')
        self.assertEqual(response.context['classified_run'], run)
        self.assertEqual(len(response.context['classified_rows']), PAGE_SIZE)
        self.assertEqual(response.context['classified_rows'][0]['row_number'], 1)
        self.assertEqual(response.context['classified_page'].paginator.count, 30)

    def test_second_page_contains_original_row_number_26(self):
        dataset = self.create_dataset(55)
        train_kmeans(dataset, ['x', 'y'], 2)

        response = self.client.get(
            reverse('dashboard:index'),
            {
                'classified_algorithm': 'kmeans',
                'classified_page': 2,
            },
        )

        self.assertEqual(len(response.context['classified_rows']), 25)
        self.assertEqual(response.context['classified_rows'][0]['row_number'], 26)
        self.assertContains(
            response,
            'classified_page=3#classified-pane',
        )

    def test_selector_lists_both_trained_algorithms(self):
        dataset = self.create_clustered_dataset()
        train_kmeans(dataset, ['x', 'y'], 2)
        train_dbscan(dataset, ['x', 'y'], epsilon=0.1, min_samples=2)

        response = self.client.get(
            reverse('dashboard:index'),
            {'classified_algorithm': 'dbscan'},
        )

        options = response.context['classified_available_algorithms']
        self.assertEqual(
            options,
            [
                {'value': 'kmeans', 'label': 'K-Means'},
                {'value': 'dbscan', 'label': 'DBSCAN'},
            ],
        )
        self.assertEqual(response.context['classified_algorithm'], 'dbscan')
        self.assertContains(response, 'Cluster DBSCAN')

    def test_dbscan_filter_includes_noise(self):
        dataset = self.create_clustered_dataset()
        run = train_dbscan(
            dataset,
            ['x', 'y'],
            epsilon=0.1,
            min_samples=2,
        )
        self.assertEqual(run.noise_count, 1)

        context = build_classified_context(
            dataset,
            requested_algorithm='dbscan',
            requested_cluster='-1',
        )

        self.assertEqual(context['classified_page'].paginator.count, 1)
        self.assertEqual(context['classified_rows'][0]['row_number'], 7)
        self.assertEqual(context['classified_rows'][0]['cluster_label'], 'Ruido')

    def test_category_training_only_exposes_analyzed_rows(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria', 'x', 'y'],
            records=[
                {'categoria': 'Agua', 'x': '0', 'y': '0'},
                {'categoria': 'Tierra', 'x': '100', 'y': '100'},
                {'categoria': 'Agua', 'x': '1', 'y': '1'},
                {'categoria': 'Tierra', 'x': '101', 'y': '101'},
                {'categoria': 'Agua', 'x': '10', 'y': '10'},
                {'categoria': 'Agua', 'x': '11', 'y': '11'},
            ],
        )
        train_kmeans(
            dataset,
            ['x', 'y'],
            2,
            requested_category='agua',
        )

        context = build_classified_context(
            dataset,
            requested_algorithm='kmeans',
        )

        self.assertEqual(
            [row['row_number'] for row in context['classified_rows']],
            [1, 3, 5, 6],
        )
        self.assertTrue(
            all(row['values'][0] == 'Agua' for row in context['classified_rows'])
        )

    def test_excel_download_matches_selected_cluster(self):
        dataset = self.create_dataset(30)
        run = train_kmeans(dataset, ['x', 'y'], 2)
        selected_cluster = str(run.assignments[0]['cluster'])
        expected_assignments = [
            assignment
            for assignment in run.assignments
            if str(assignment['cluster']) == selected_cluster
        ]

        response = self.client.get(
            reverse('classified_data:download'),
            {'algorithm': 'kmeans', 'cluster': selected_cluster},
        )
        workbook = load_workbook(BytesIO(response.content), read_only=False)
        worksheet = workbook['Datos clasificados']
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual(response.status_code, 200)
        self.assertIn('datos-clasificados-kmeans', response['Content-Disposition'])
        self.assertEqual(
            list(rows[0]),
            ['Número de fila original', 'Cluster K-Means', 'nombre', 'x', 'y'],
        )
        self.assertEqual(len(rows) - 1, len(expected_assignments))
        self.assertEqual(
            [row[0] for row in rows[1:]],
            [assignment['row_number'] for assignment in expected_assignments],
        )
        self.assertTrue(
            all(row[1] == f'Cluster {selected_cluster}' for row in rows[1:])
        )
        self.assertIn('DatosClasificados', worksheet.tables)

    def test_download_rejects_algorithm_without_training(self):
        self.create_dataset(10)

        response = self.client.get(
            reverse('classified_data:download'),
            {'algorithm': 'dbscan'},
        )

        self.assertEqual(response.status_code, 404)

    def test_download_only_contains_rows_analyzed_by_category(self):
        dataset = Dataset.objects.create(
            pk=1,
            source_name='categorias.csv',
            columns=['categoria', 'x'],
            records=[
                {'categoria': 'Agua', 'x': '0'},
                {'categoria': 'Tierra', 'x': '100'},
                {'categoria': 'Agua', 'x': '1'},
                {'categoria': 'Tierra', 'x': '101'},
                {'categoria': 'Agua', 'x': '10'},
                {'categoria': 'Agua', 'x': '11'},
            ],
        )
        train_kmeans(
            dataset,
            ['x'],
            2,
            requested_category='agua',
        )

        response = self.client.get(
            reverse('classified_data:download'),
            {'algorithm': 'kmeans'},
        )
        worksheet = load_workbook(
            BytesIO(response.content),
            read_only=True,
        )['Datos clasificados']
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual([row[0] for row in rows[1:]], [1, 3, 5, 6])
        self.assertTrue(all(row[2] == 'Agua' for row in rows[1:]))

    def test_download_rejects_invalid_cluster_filter(self):
        dataset = self.create_dataset(10)
        train_kmeans(dataset, ['x', 'y'], 2)

        response = self.client.get(
            reverse('classified_data:download'),
            {'algorithm': 'kmeans', 'cluster': '999'},
        )

        self.assertEqual(response.status_code, 404)

    def test_dbscan_results_no_longer_repeat_assignment_table(self):
        dataset = self.create_clustered_dataset()
        train_dbscan(dataset, ['x', 'y'], epsilon=0.1, min_samples=2)

        response = self.client.get(
            reverse('dashboard:index'),
            {'results_view': 'dbscan'},
        )

        self.assertNotContains(response, 'Asignaciones por registro')
        self.assertNotContains(response, 'Páginas de resultados DBSCAN')

    @staticmethod
    def create_dataset(row_count):
        return Dataset.objects.create(
            pk=1,
            source_name='personas.csv',
            columns=['nombre', 'x', 'y'],
            records=[
                {
                    'nombre': f'Persona {index}',
                    'x': str(index),
                    'y': str(index % 5),
                }
                for index in range(row_count)
            ],
        )

    @staticmethod
    def create_clustered_dataset():
        return Dataset.objects.create(
            pk=1,
            source_name='grupos.csv',
            columns=['x', 'y'],
            records=[
                {'x': '0', 'y': '0'},
                {'x': '0.05', 'y': '0.04'},
                {'x': '0.1', 'y': '0'},
                {'x': '10', 'y': '10'},
                {'x': '10.05', 'y': '10.04'},
                {'x': '10.1', 'y': '10'},
                {'x': '30', 'y': '30'},
            ],
        )
